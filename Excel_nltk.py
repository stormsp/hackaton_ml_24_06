import nltk
import openpyxl
import pickle
from nltk.tokenize import sent_tokenize


# Загрузка классификатора
with open('classifier.pickle', 'rb') as file:
    classifier = pickle.load(file)

# Подключение excel
wb = openpyxl.load_workbook("Датасет.xlsx")
ws = wb["Лист1"]


def extract_features(words):
    features = {word: True for word in words}
    return features

#  основная функция обработки
def processing(text):
    req = ''
    con = ''
    sk = ''

    #
    sentences = sent_tokenize(text)
    for sentence in sentences:
        words = nltk.word_tokenize(sentence)
        featureset = extract_features(words)
        label = classifier.classify(featureset)

        if label == 'требования к сотруднику':
            req += sentence + '\n'
        elif label == 'условия работы':
            con += sentence + '\n'
        elif label == 'персональные навыки':
            sk += sentence + '\n'
    return [req, con, sk]


#  ws.max_row
mx = 1000
for row in range(2, mx + 1):
    #  получение ячеек
    responsibilities = ws.cell(row, 4)  # Должностные обязанности
    requirements = ws.cell(row, 5)  # Требования к соискателю
    terms = ws.cell(row, 6)  # Условия
    skills = ws.cell(row, 7)  # Ключевые навыки

    #  очистка ячеек
    requirements.value = ''
    terms.value = ''
    skills.value = ''

    # обработка
    result = processing(responsibilities.value)
    requirements_res = result[0]
    terms_res = result[1]
    skills_res = result[2]

    # запись
    requirements.value = requirements_res
    terms.value = terms_res
    skills.value = skills_res

    # предотвращает от ошибок
    requirements.data_type = 'str'
    terms.data_type = 'str'
    skills.data_type = 'str'

    print('строка {0}'.format(row))
    print_mode = False
    if print_mode:
        # print('Должностные обязанности: ', '\n', responsibilities.value, '\n')
        print('--требования к сотруднику--: ', '\n', requirements_res)
        print('--условия работы--: ', '\n', terms_res)
        print('--персональные навыки--: ', '\n', '\n', skills_res)


wb.save("Отредактировано.xlsx")